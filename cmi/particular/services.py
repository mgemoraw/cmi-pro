import csv
from io import TextIOWrapper
# import pandas as pd
from openpyxl import load_workbook
from . models import Particular, Task, Division, ProjectType


class GridDataParser:
    def __init__(self, file, file_format):
        self.file = file
        self.format = file_format.lower()

    def parse(self):
        """Main method to parse the file based on its format."""
        if self.format == 'csv':
            return self.parse_csv()
        elif self.format in ('xlsx', 'xls'):
            return self.parse_xlsx()
        else:
            raise ValueError(f"Unsupported file format: {self.format}")

    def parse_csv(self, delimiter=","):
        """Parse CSV file into a list of rows (lists)."""
        try:
            # Try UTF-8 first
            file_data = TextIOWrapper(self.file.file, encoding='utf-8', errors='replace')
            reader = csv.reader(file_data, delimiter=delimiter)
            rows = [row for row in reader]
            headers = rows[0]
            data_rows = rows[1:]
            for row in data_rows:
        
                subsector = ProjectType.objects.get(name=row[2])
                # division = Division.objects.filter(name=row[4]).first()
                # if not(division.exists):
                #     division = Division.objects.create(
                #         code=row[3],
                #         name=row[4],
                #     )
                division, div_created = Division.objects.get_or_create(
                    code=row[3],
                    name=row[4],
                )

                # check if particular already registered
                particular = Particular.objects.filter(
                    pid=row[1],
                ).exists()
                if particular:
                    continue
                else:
                    particular, created = Particular.objects.get_or_create(
                        pid=row[1],
                        project_type=subsector,
                        division=division,
                        task=row[5].strip(),
                        element=row[6].strip(),
                        name=row[7].strip(),
                    )
                    
            return rows  

        except Exception as e:
            raise ValueError(f"Error parsing CSV: {e}")
        return rows

    def parse_xlsx(self):
        """Parse XLSX file into a list of rows (lists)."""
        try:
            wb = load_workbook(self.file.file, data_only=True)
            sheet = wb.active  # You can customize which sheet to use
            rows = [[cell for cell in row] for row in sheet.iter_rows(values_only=True)]
            print(rows)
        except Exception as e:
            raise ValueError(f"Error parsing XLSX: {e}")
        return rows
