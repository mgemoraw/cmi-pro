from openpyxl import Workbook, load_workbook
from django.conf import settings
from django.core.files.uploadedfile import InMemoryUploadedFile
from io import BytesIO, StringIO
import os
import csv
from core.models import Project, Engineer, Collector


class InstanceDataParser:
    def __init__(self, data_file: InMemoryUploadedFile, data_type=None, format=None, *args, **kwargs):
        self.data_file = data_file
        self.data_type = data_type
        self.format = format

        if self.kwargs:
            self.extension = kwargs.get('extension', None)
            self.method = kwargs.get('method', None)

    def parse(self):
        if self.data_type =="project":
            pass
        elif self.data_type == "instance":
            self._parse_data_instance()

        elif self.data_type == "particular":
            pass
        elif self.data_type == "engineer":
            pass
        elif self.data_type == "collector":
            pass
        else:
            raise ValueError("Data Type is Unknown!")
    

    def _parse_data_instance(self, method=None):
        work_methods = ["truck", "dozer", "excavator", "tipper", 'labor', "work sampling", 'project']

        if self.method != None:
            if self.method == "truck":
                self._parse_truck_instance()
            elif self.method == "excavator":
                self._parse_excavator_instance()
            elif self.method == "labor":
                self._parse_labor_instance()
            elif self.method == "dozer":
                self._parse_dozer_instance()
            elif self.method == "work sampling":
                self._parse_ws_instance()
            
            else:
                raise ValueError("Work Method {} is not valid".format(self.method))
        
        else:
            raise ValueError("Work methodlogy (equipment) not provided")

    def _parse_truck_instance(self):
        # self.format = self.format.lower()
        if self.format is not None:
            try:
                if self.format == 'xlsx':
                    self._parse_truck_xlsx()
                    
            except Exception as e:
                raise e
            

    def _parse_truck_xlsx(self):
        wb = load_workbook(filename=BytesIO(self.data_file.read()), data_only=True)
        ws = wb.active

        headers = [cell.value for cell in ws[1]]
        expected_headers = [
            'project', 'collector', 'particular', 'date', 'tipper', 'task', 'number_of_equipment_types',
            'cycle_number', 'manpwer', 'soil_type', 'unit', 'load_cycle', 'haul_cyle', 'dump_cycle',
            'return_cycle', 'total_cycle', 'q_heaped_capacity', 'productivity'
        ]

    def _parse_labor_instance(self):
        pass 

    def _parse_excavator_instance(self):
        pass 
    
    def _parse_dozer_instance(self):
        pass 
    
    def _parse_ws_instance(self):
        pass 
    