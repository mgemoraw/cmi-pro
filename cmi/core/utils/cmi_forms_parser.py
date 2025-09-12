from openpyxl import load_workbook

class LaborFormParser:
    def __init__(self, file):
        self.wb = load_workbook(file, data_only=True)
        self.labor_ws = None
        self.ws_ws = None
        self.dv_ws = None

        try:
            self.labor_ws = self.wb['labour']
            self.ws_ws = self.wb['work_sampling']
            self.dv_ws = self.wb['daily_variables']
        except KeyError as e:
            print(f"Error: Sheet '{e.args[0]}' not found in the workbook.")
            # Handle the error or re-raise
            raise
        
    def save(self, obj):
        pass


    def parse(self):
        """
        Parses all three sheets: labour, work_sampling, and daily_variables.
        Returns a dictionary containing the parsed data from each sheet.
        """
        parsed_data = {
            'labour_data': self._parse_labour_data(),
            'work_sampling_data': self._parse_work_sampling_data(),
            'daily_variables_data': self._parse_daily_variables_data()
        }
        return parsed_data

    def _parse_labour_data(self):
        """
        Parses the 'labour' sheet to extract key labor productivity data.
        """
        if not self.labor_ws:
            return None

        # Headers are on row 17, and data starts on row 18
        data_start_row = 8
        data = []
        # for col in self.labor_ws.iter_cols(min_row=6, max_col=10, max_row=7):
        #     for cell in col:
        #         print(f"{cell}: {cell.value}")


        # Iterate over rows starting from the data_start_row
        for row in self.labor_ws.iter_rows(min_row=data_start_row, values_only=True):
            # Check if the row is empty (based on a key column like 'Date')
            if not row[1]:
                # print("No rows")
                continue
            
            row_data = {
                'Data Count': row[0],
                'Date': row[1],
                'Project Code': row[2],
                'Data Collector': row[3],
                'Number of crews': row[4],
                'Task Type': row[5],
                'Particulars': row[6],
                'Location': row[7],
                'Crew Size': row[8],
                'Crew Composition': row[9],
                'Work Time - R': row[10],
                'Work Time - OT': row[11],
                'Total Work Time': row[12],
                'Total Manhours': row[13],
                'Unit': row[14],
                'Daily Units Completed': row[15],
                'Productivity': row[16],
                'Problem Code': row[17]
            }
            data.append(row_data)
        return data

    def _parse_work_sampling_data(self):
        """
        Parses the 'work_sampling' sheet to extract the total sum for each observation.
        """
        if not self.ws_ws:
            return None
        
        # The data structure is complex, with observation numbers across two sections.
        # We need to find all rows containing 'Sum'
        sums_data = []
        for row in self.ws_ws.iter_rows(values_only=True):
            # The 'Sum' label is in the first column
            if row[0] == 'Sum':
                # The sums are in the columns starting from the 5th column
                # The data is repeated for each crew observation
                sums = [value for value in row[4:] if value is not None and isinstance(value, (int, float))]
                if sums:
                    sums_data.append(sums)
        
        return sums_data

    def _parse_daily_variables_data(self):
        """
        Parses the 'daily_variables' sheet to extract all variables and their descriptions.
        """
        if not self.dv_ws:
            return None

        variables = []
        # Data starts from row 6, and the variable name is in column 3
        # and description in column 6. The ID is in column 2.
        for row in self.dv_ws.iter_rows(min_row=6, values_only=True):
            # Check if the row has a variable ID to skip empty rows
            if row[1]:
                variable_data = {
                    'Factor ID': row[1],
                    'Sub-Factors': row[2],
                    'Scale of Measure': row[3],
                    'Range of Value': row[4],
                    'Description': row[5],
                    'Value': row[6]
                }
                variables.append(variable_data)
        
        return variables

if __name__ == '__main__':
    # You would replace 'labor_records.xlsx' with your actual file path
    try:
        parser = LaborFormParser('labor_records.xlsx')
        parsed_results = parser.parse()
        print(parsed_results.get('work_sampling_data'))
        # print(parsed_results.get('work_sampling_data'))
        # print(parsed_results.get('daily_variables_data'))
        # You can now print the results to verify
        # for sheet_name, data in parsed_results.items():
        #     print(f"--- {sheet_name.replace('_', ' ').title()} ---")
        #     # print(data) # To see the full data structure
        #     if data and isinstance(data, list):
        #         print(f"Found {len(data)} records.")
        #     else:
        #         print("No data found or data is not a list.")

    except Exception as e:
        print(f"An error occurred: {e}")