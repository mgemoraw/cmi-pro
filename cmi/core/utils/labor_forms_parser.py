from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, get_column_interval
from datetime import datetime, timedelta, time
import re


class LaborFormParser:
    def __init__(self, file):
        self.wb = load_workbook(file, data_only=True)
        self.labor_ws = None
        self.ws_ws = None
        self.dv_ws = None

        try:
            self.labor_ws = self.wb['labour']
            self.ws_ws = self.wb['work_sampling']
            self.dv_ws = self.wb['daily_variables']
        except KeyError as e:
            print(f"Error: Sheet '{e.args[0]}' not found in the workbook.")
            # Handle the error or re-raise
            raise
        
    def save(self, obj):
        pass


    def parse(self):
        """
        Parses all three sheets: labour, work_sampling, and daily_variables.
        Returns a dictionary containing the parsed data from each sheet.
        """
        parsed_data = {
            'labour': self._parse_labour_data(),
            'ws': self._parse_work_sampling_data(),
            'dv': self._parse_daily_variables_data()
        }
        return parsed_data

    def _parse_labour_data(self):
        """
        Parses the 'labour' sheet to extract key labor productivity data.
        """
        if not self.labor_ws:
            return None

        # Headers are on row 17, and data starts on row 18
        data_start_row = 8
        data = []
        # for col in self.labor_ws.iter_cols(min_row=6, max_col=10, max_row=7):
        #     for cell in col:
        #         print(f"{cell}: {cell.value}")


        # Iterate over rows starting from the data_start_row
        for row in self.labor_ws.iter_rows(min_row=data_start_row, values_only=True):
            # Check if the row is empty (based on a key column like 'Date')
            if not row[1]:
                # print("No rows")
                continue
            
            row_data = {
                'Data Count': row[0],
                'Date': row[1],
                'Project Code': row[2],
                'Data Collector': row[3],
                'Number of crews': row[4],
                'Task Type': row[5],
                'Particulars': row[6],
                'Location': row[7],
                'Crew Size': row[8],
                'Crew Composition': row[9],
                'Work Time - R': row[10],
                'Work Time - OT': row[11],
                'Total Work Time': row[12],
                'Total Manhours': row[13],
                'Unit': row[14],
                'Daily Units Completed': row[15],
                'Productivity': row[16],
                'Problem Code': row[17]
            }
            data.append(row_data)
        return data

    def _parse_work_sampling_data(self):
        """
        Parses the 'work_sampling' sheet to extract the total sum for each observation.
        """
        if not self.ws_ws:
            return None
        
        # The data structure is complex, with observation numbers across two sections.
        # We need to find all rows containing 'Sum'
        ws_data = {}

        observations = []
        counter = 0
        data_date = self.ws_ws['C10'].value
        location = self.ws_ws['B13'].value
        task_type = self.ws_ws['B15'].value
        project_code = self.ws_ws['B11'].value
        data_collector = f"{self.ws_ws['B21'].value} {self.ws_ws['C21'].value}"

        ws_data['date'] = data_date
        ws_data['location'] = location
        ws_data['task type'] = task_type
        ws_data['project code'] = project_code 
        ws_data['data collector'] = data_collector

        # print("########### project ################")
        # print(f"## Project Code: {project_code}")
        # print(f"## Date: {data_date}")
        # print(f"## location: {location}")
        # print(f"## task type: {task_type}")
        # print(f"## Data collector: {data_collector}")

        for r in range(12, 60, 12):
            oh = None # instatiate observation hour to
            for i in range(0, 36):
                COL = get_column_letter(i+6)
                # print(COL)
                cell = f"{COL}{r-3}"
                
                # read every cell of the row that could contain the observation hour
                observation_hour = self.ws_ws[cell].value
                if observation_hour is not None:
                    oh = observation_hour
                
                observation_minute = self.ws_ws[f"{COL}{11+r-12}"].value
                observation_time = self._get_observation_time(oh, observation_minute)

                if observation_minute is not None:
                    observation = {
                        # "observation_hour": oh,
                        "observation number": self.ws_ws[f"{COL}{10+r-12}"].value,
                        "observation time": observation_time ,
                        "direct": self.ws_ws[f"{COL}{12+r-12}"].value,
                        "preparatory": self.ws_ws[f"{COL}{13+r-12}"].value, 
                        "tools and equipment": self.ws_ws[f"{COL}{r+14-12}"].value, 
                        "Material handling": self.ws_ws[f"{COL}{r+15-12}"].value, 
                        "Waiting": self.ws_ws[f"{COL}{r+16-12}"].value,
                        "Travel": self.ws_ws[f"{COL}{r+17-12}"].value,
                        "Personal": self.ws_ws[f"{COL}{r+18-12}"].value,
                        "sum": self.ws_ws[f'{COL}{r+19-12}'].value,
                    }
                    observations.append(observation)

        ws_data['observations'] = observations
        

        for row in self.ws_ws.iter_rows(values_only=True):
            """parse work sampling general data first"""
            try:
                cols = []
               
                # for cell in row:

                #     if cell:
                #         if (str(cell).lower().startswith('observation time')):
                #             counter += 1
                #             observations.append('observation: {}'.format(counter))
                #             # observations
                #         cols.append(cell)
                # print(cols if len(cols) != 0 else "")
            except Exception as e:
                raise e
             
            
        
        return ws_data

    def _get_observation_time(self, oh: str, om: int, use_end: bool = False) -> datetime.time:
        """
        oh: hour range string like "2:00:00 - 3:00:00"
        om: minute (0-59)
        use_end: if True, use the ending hour; otherwise use starting hour

        """

        try:
            if oh is not None and om is not None:
                # print(oh)
                stime, etime = oh.split("-")
                # print(f"start: {stime}, end: {etime}")
                sh = oh.split(':')[0]
                eh = oh.split(':')[0]

                if use_end:
                    hour = sh   # "3:00:00"
                else:
                    hour = eh   # "2:00:00"
                
                # split into components
                # h, m, s = map(int, time_part.split(':'))
        
                return time(int(hour), int(om), 0)
        except Exception as e:
            pass 
            raise ValueError("hour and minute cannot be None")


    def _parse_daily_variables_data(self):
        """
        Parses the 'daily_variables' sheet to extract all variables and their descriptions.
        """
        if not self.dv_ws:
            return None

        dv_data = {}
        data_date = self.dv_ws['H6'].value
        task_type = self.dv_ws['E5'].value
        project_code = self.dv_ws['C5'].value
        data_collector = f"{self.dv_ws['C6'].value}"

        dv_data['date'] = data_date
        dv_data['task type'] = task_type
        dv_data['project code'] = project_code 
        dv_data['data collector'] = data_collector
        variables = []
        # Data starts from row 6, and the variable name is in column 3
        # and description in column 6. The ID is in column 2.
        for row in self.dv_ws.iter_rows(min_row=6, values_only=True):
            # Check if the row has a variable ID to skip empty rows
            if row[1]:
                variable_data = {
                    'Factor ID': row[1],
                    'Sub-Factors': row[2],
                    'Scale of Measure': row[3],
                    "data source": row[4],
                    'Data Value': row[5],
                    'Range of Value': row[6],
                    'Description': row[7],

                }
                variables.append(variable_data)
        
        dv_data['variables'] = variables

        return dv_data
    
    

if __name__ == '__main__':
    # You would replace 'labor_records.xlsx' with your actual file path
    try:
        parser = LaborFormParser('05_05_2017_gedefaw labor_records_form.xlsx')
        parsed_results = parser.parse()
        print(parsed_results.get('ws'))
        # print(parsed_results.get('labour'))
        print(parsed_results.get('dv'))
        # You can now print the results to verify
        # for sheet_name, data in parsed_results.items():
        #     print(f"--- {sheet_name.replace('_', ' ').title()} ---")
        #     # print(data) # To see the full data structure
        #     if data and isinstance(data, list):
        #         print(f"Found {len(data)} records.")
        #     else:
        #         print("No data found or data is not a list.")

    except Exception as e:
        print(f"An error occurred: {e}")