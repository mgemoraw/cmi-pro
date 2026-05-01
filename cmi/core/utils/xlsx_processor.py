import os 
import zipfile
from openpyxl import load_workbook

import argparse
from pathlib import Path

def get_args():
    parser = argparse.ArgumentParser(description="Process Excel files")

    parser.add_argument(
        "--input",
        type=str,
        required=True,
        help="Input folder containing Excel files"
    )

    parser.add_argument(
        "--output",
        type=str,
        default="outputs",
        help="Output folder"
    )

    parser.add_argument(
        "--template",
        type=str,
        required=True,
        help="Path to target template"
    )

    return parser.parse_args()


def get_excel_files(input_dir):
    path = Path(input_dir)

    if not path.exists():
        raise FileNotFoundError(f"{input_dir} does not exist")

    # Get all Excel files
    files = list(path.glob("*.xlsx")) + list(path.glob("*.xls"))

    if not files:
        raise ValueError("No Excel files found")

    return files


class XLSXProcessor:
    def __init__(self, template_path: str, config):
        self.template_path = template_path
        self.config = config

    def load_template(self):
        wb = load_workbook(self.template_path)
        ws = wb.active
        return wb, ws

    # -------------------------------
    # HANDLE MERGED CELLS
    # -------------------------------
    def _get_value(self, ws, cell):
        for merged in ws.merged_cells.ranges:
            if cell.coordinate in merged:
                return ws.cell(merged.min_row, merged.min_col).value
        return cell.value

    # -------------------------------
    # EXTRACT DATA FROM SOURCE FILE
    # ------------------------------- 
    def extract(self, file_path: str):
        wb = load_workbook(file_path, data_only=True)
        ws = wb.active

        data = []
        start_row = self.config['source']['start_row']
        col_map = self.config["source"]["columns"]

        for row in ws.iter_rows(min_row=start_row):
            if not row[col_map['cycle'] - 1].value:  # Assuming 'cycle' is a mandatory field
                continue

            record = {}
            for key, col_index in col_map.items():
                cell = row[col_index - 1]
                record[key] = self._get_value(ws, cell)
            data.append(record)
        return data


    # -------------------------------
    # SPLIT DATA INTO CHUNKS
    # -------------------------------
    def chunk(self, data):
        size = self.config["chunk_size"]
        for i in range(0, len(data), size):
            yield data[i:i + size]
        

    # -------------------------------
    # WRITE INTO TEMPLATE
    # -------------------------------
    def write(self, chunk, metadata=None):
        wb = load_workbook(self.template_path)
        ws = wb.active

        # write metadata
        if metadata:
            for key, cell in self.config.get("metadata", {}).items():
                ws[cell] = metadata.get(key)

        start_row = self.config['target']['start_row']
        col_map = self.config["target"]["columns"]

        for i, row in enumerate(chunk):
            r = start_row + i 

            for key, col_letter in col_map.items():
                ws[f"{col_letter}{r}"] = row.get(key, "")

        return wb
    

    # -------------------------------
    # PROCESS FILES + GENERATE OUTPUTS
    # -------------------------------
    def process(self, source_files: list, output_dir: str):
        os.makedirs(output_dir, exist_ok=True)
        output_files = []

        counter = 1

        for file in source_files:
            data = self.extract(file)

            for chunk in self.chunk(data):
                wb = self.write(chunk)
                output_path = os.path.join(output_dir, f"output_{counter}.xlsx")
                wb.save(output_path)
                output_files.append(output_path)
                counter += 1

        return output_files
    

    # -------------------------------
    # ZIP OUTPUT FILES
    # -------------------------------
    def zip_outputs(self, files, zip_name="outputs.zip"):
        with zipfile.ZipFile(zip_name, 'w') as z:
            for f in files:
                z.write(f, os.path.basename(f))
        return zip_name
          

