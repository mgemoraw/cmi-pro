from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, get_column_interval
from datetime import datetime, timedelta, time
import re


class EquipmentFormsParser:
    def __init__(self, file):
        self.wb = load_workbook(file, data_only=True)
        self.equipment_ws = None
        self.mpdm_ws = None
        self.dv_ws = None
        self.equipment_type = self._map_equipment_form()

        try:
            self.equipment_ws = self.wb[self.equipment_type]
            self.mpdm_ws = self.wb['mpdm']
            self.dv_ws = self.wb['daily_variables']
        except KeyError as e:
            print(f"Error: Sheet '{e.args[0]}' not found in the workbook.")
            # Handle the error or re-raise
            raise
    
    def _map_equipment_form(self):
        sheet_names = self.wb.sheetnames
        equipments = [
            "truck",
            "dozer",
            "excavator",
            "grader",
            "asphalt_paver",
            "roller",
        ]
        if sheet_names[0] and sheet_names[0] in equipments:
            return sheet_names[0].lower()
        
    def parse(self):
        """
        Parses all three sheets: labour, work_sampling, and daily_variables.
        Returns a dictionary containing the parsed data from each sheet.
        """
        equipment_data = None
        if self.equipment_type is not None:
            if self.equipment_type == 'truck':
                equipment_data = self._parse_truck_data()
            elif self.equipment_type == 'excavator':
                equipment_data = self._parse_excavator_data()
            elif self.equipment_type == 'dozer':
                equipment_data = self._parse_dozer_data()
            else:
                raise ValueError("Equipment Form Type Unknown")

        parsed_data = {
            self.equipment_type: equipment_data,
            'mpdm': self._parse_mpdm_data(),
            'dv': self._parse_daily_variables_data()
        }
        return parsed_data


    def _parse_truck_data(self):
        """
        Parses the 'labour' sheet to extract key labor productivity data.
        """
        if not self.equipment_ws:
            return None

        # Headers are on row 6, and data starts on row 7
        data_start_row = 7
        data = []
        # for col in self.labor_ws.iter_cols(min_row=6, max_col=10, max_row=7):
        #     for cell in col:
        #         print(f"{cell}: {cell.value}")


        # Iterate over rows starting from the data_start_row
        counter = 0
        for row in self.equipment_ws.iter_rows(min_row=data_start_row, values_only=True):
            # Check if the row is empty (based on a key column like 'Date')
            if not row[1]:
                # print("No rows")
                continue
            row_index = data_start_row + counter
            row_data = {
                'date': self.equipment_ws[f'A{row_index}'].value,
                'Project Code':self.equipment_ws[f'B{row_index}'].value ,
                'Data Collector': self.equipment_ws[f'C{row_index}'].value,
                "number of equipment types": self.equipment_ws[f'D{row_index}'].value,
                "equipment tag": self.equipment_ws[f'E{row_index}'].value,
                "manpower": self.equipment_ws[f'F{row_index}'].value,
                "truck plate": self.equipment_ws[f'G{row_index}'].value,
                "task type": self.equipment_ws[f'H{row_index}'].value,
                "description": self.equipment_ws[f"I{row_index}"].value,
                "soil type": self.equipment_ws[f"J{row_index}"].value,
                "total cycle time": self.equipment_ws[f"O{row_index}"].value,
                'unit of time': self.equipment_ws[f"P{row_index}"].value,
                'actual bucket capacity': self.equipment_ws[f"Q{row_index}"].value,
                'productivity': self.equipment_ws[f"R{row_index}"].value,
            }
            data.append(row_data)
            counter += 1
        return data

    def _parse_dozer_data(self):
        """
        Parses the 'labour' sheet to extract key labor productivity data.
        """
        if not self.equipment_ws:
            return None

        # Headers are on row 6, and data starts on row 7
        data_start_row = 7
        data = []
        # for col in self.labor_ws.iter_cols(min_row=6, max_col=10, max_row=7):
        #     for cell in col:
        #         print(f"{cell}: {cell.value}")


        # Iterate over rows starting from the data_start_row
        counter = 0
        for row in self.equipment_ws.iter_rows(min_row=data_start_row, values_only=True):
            # Check if the row is empty (based on a key column like 'Date')
            if not row[4] or not row[15]:
                # print("No rows")
                continue
            row_index = data_start_row + counter
            row_data = {
                'date': self.equipment_ws[f'A{row_index}'].value,
                'Project Code':self.equipment_ws[f'B{row_index}'].value ,
                'Data Collector': self.equipment_ws[f'C{row_index}'].value,
                "number of equipment types": self.equipment_ws[f'D{row_index}'].value,
                "dozer cycle": self.equipment_ws[f'E{row_index}'].value,
                "manpower": self.equipment_ws[f'F{row_index}'].value,
                "dozer blade type": self.equipment_ws[f'G{row_index}'].value,
                "task type": self.equipment_ws[f'H{row_index}'].value,
                "description": self.equipment_ws[f"I{row_index}"].value,
                "soil type": self.equipment_ws[f"J{row_index}"].value,
                "blade height": self.equipment_ws[f"K{row_index}"].value,
                "blade width": self.equipment_ws[f"L{row_index}"].value,
                "blade length": self.equipment_ws[f"M{row_index}"].value,
                'unit of time': self.equipment_ws[f"N{row_index}"].value,
                'blade load': self.equipment_ws[f"O{row_index}"].value,
                "total cycle time": self.equipment_ws[f"P{row_index}"].value,
                'productivity': self.equipment_ws[f"Q{row_index}"].value,
            }
            data.append(row_data)
            counter += 1
        return data

    def _parse_excavator_data(self):
        """
        Parses the 'labour' sheet to extract key labor productivity data.
        """
        if not self.equipment_ws:
            return None

        # Headers are on row 6, and data starts on row 7
        data_start_row = 7
        data = []
        # for col in self.labor_ws.iter_cols(min_row=6, max_col=10, max_row=7):
        #     for cell in col:
        #         print(f"{cell}: {cell.value}")


        # Iterate over rows starting from the data_start_row
        counter = 0
        for row in self.equipment_ws.iter_rows(min_row=data_start_row, values_only=True):
            # Check if the row is empty (based on a key column like 'Date')
            if not row[4] or not row[17]:
                # print("No rows")
                continue
            row_index = data_start_row + counter
            row_data = {
                'date': self.equipment_ws[f'A{row_index}'].value,
                'Project Code':self.equipment_ws[f'B{row_index}'].value ,
                'Data Collector': self.equipment_ws[f'C{row_index}'].value,
                "number of equipment types": self.equipment_ws[f'D{row_index}'].value,
                "excavator cycle": self.equipment_ws[f'E{row_index}'].value,
                "manpower": self.equipment_ws[f'F{row_index}'].value,
                "excavator type": self.equipment_ws[f'G{row_index}'].value,
                "task type": self.equipment_ws[f'H{row_index}'].value,
                "description": self.equipment_ws[f"I{row_index}"].value,
                "soil type": self.equipment_ws[f"J{row_index}"].value,
                "bucket fill factor": self.equipment_ws[f"K{row_index}"].value,
                "angle of swing": self.equipment_ws[f"L{row_index}"].value,
                "depth of cut": self.equipment_ws[f"M{row_index}"].value,
                'volume correction': self.equipment_ws[f"N{row_index}"].value,
                'efficiency': self.equipment_ws[f"O{row_index}"].value,
                'unit of time': self.equipment_ws[f"P{row_index}"].value,
                'heaped bucket capacity': self.equipment_ws[f"Q{row_index}"].value,
                "total cycle time": self.equipment_ws[f"R{row_index}"].value,
                'productivity': self.equipment_ws[f"S{row_index}"].value,
            }
            data.append(row_data)
            counter += 1
        return data
 


    def _parse_mpdm_data(self):
        """
        Parses the 'labour' sheet to extract key labor productivity data.
        """
        if not self.mpdm_ws:
            return None

        # Headers are on row 6, and data starts on row 7
        data_start_row = 7
        data = []
        # for col in self.labor_ws.iter_cols(min_row=6, max_col=10, max_row=7):
        #     for cell in col:
        #         print(f"{cell}: {cell.value}")


        # Iterate over rows starting from the data_start_row
        counter = 0
        for row in self.mpdm_ws.iter_rows(min_row=data_start_row, values_only=True):
            # Check if the row is empty (based on a key column like 'Date')
            if not row[7]:
                # print("No rows")
                # print(row)
                continue
            row_index = data_start_row + counter
            row_data = {
                'date': self.mpdm_ws[f'A{row_index}'].value,
                'Data Collector': self.mpdm_ws[f'B{row_index}'].value,
                'unit of time': self.mpdm_ws[f"C{row_index}"].value,
                'Project Code':self.mpdm_ws[f'D{row_index}'].value ,
                
                'operation': self.mpdm_ws[f"E{row_index}"].value,
                'equipment': self.mpdm_ws[f"F{row_index}"].value,
                'production cycle number': self.mpdm_ws[f"G{row_index}"].value,
                'cycle time': self.mpdm_ws[f"H{row_index}"].value,
                'environment': self.mpdm_ws[f"I{row_index}"].value,
                'equipment': self.mpdm_ws[f"J{row_index}"].value,
                'labour': self.mpdm_ws[f"K{row_index}"].value,
                'material': self.mpdm_ws[f"L{row_index}"].value,
                'management': self.mpdm_ws[f"M{row_index}"].value,
                'others': self.mpdm_ws[f"N{row_index}"].value,
            }
            data.append(row_data)
            counter += 1
        return data

    def _parse_daily_variables_data(self):
        """
        Parses the 'daily_variables' sheet to extract all variables and their descriptions.
        """
        if not self.dv_ws:
            return None

        dv_data = {}
        data_date = self.dv_ws['H6'].value
        task_type = self.dv_ws['E5'].value
        project_code = self.dv_ws['C5'].value
        data_collector = f"{self.dv_ws['C6'].value}"

        dv_data['date'] = data_date
        dv_data['task type'] = task_type
        dv_data['project code'] = project_code 
        dv_data['data collector'] = data_collector
        variables = []
        # Data starts from row 6, and the variable name is in column 3
        # and description in column 6. The ID is in column 2.
        for row in self.dv_ws.iter_rows(min_row=6, values_only=True):
            # Check if the row has a variable ID to skip empty rows
            if row[1]:
                variable_data = {
                    'Factor ID': row[1],
                    'Sub-Factors': row[2],
                    'Scale of Measure': row[3],
                    "data source": row[4],
                    'Data Value': row[5],
                    'Range of Value': row[6],
                    'Description': row[7],

                }
                variables.append(variable_data)
        
        dv_data['variables'] = variables

        return dv_data
    

if __name__ == "__main__":
    parser = EquipmentFormsParser('dozer_records.xlsx')
    data = parser.parse()

    print(f'parsing {parser.equipment_type} data...')
    print(data.get(parser.equipment_type))
    print(data.get("mpdm"))