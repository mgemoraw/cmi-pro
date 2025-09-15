from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, get_column_interval
from datetime import datetime, timedelta, time
import re


class InstanceSummaryParser:
    def __init__(self, file):
        self.wb = load_workbook(file, data_only=True)
        self.ws = None
        self.dam_ws = None
        self.irrigation_ws = None
        self.ws_ws = None
    
        try:
            self.dam_ws = self.wb["Dam"]
            self.ws_ws = self.wb['WaterSupply']
            self.irrigation_ws = self.wb['Irrigation']
        except KeyError as e:
            print(f"Error: Sheet '{e.args[0]}' not found in the workbook.")
            # Handle the error or re-raise
            raise

    def parse(self):

        return {
            'dam': self._parse_data_summary("Dam"),
            'irrigation': self._parse_data_summary("Irrigation"),
            'water_supply': self._parse_data_summary("WaterSupply"),
        }
    
    def _parse_data_summary(self, sheet_name):
        self.ws = self.wb[sheet_name]

        data = []
        data_start_row = 8
        headers = []
        # ('No', 'Date', 'Project Code', 'Data Collector', 'PID', 'Particlular', 'Problems', 'Status', 'PE', 'Status', 'Encoder')
        counter = 0
        for row in self.dam_ws.iter_rows(min_row=data_start_row, values_only=True):
            row_index = counter + data_start_row

            instance = {
                'No': self.ws[f"A{row_index}"].value,
                'Date': self.ws[f"B{row_index}"].value,
                'Project Code': self.ws[f"C{row_index}"].value,
                'Data Collector': self.ws[f"D{row_index}"].value,
                'PID': self.ws[f"E{row_index}"].value,
                'Particular': self.ws[f"F{row_index}"].value,
                'Problems': self.ws[f"G{row_index}"].value,
                'Status': self.ws[f"H{row_index}"].value,
                'PE': self.ws[f"I{row_index}"].value,
                'Encoded': self.ws[f"J{row_index}"].value,
                'Encoder': self.ws[f"K{row_index}"].value,
            }
            counter += 1
            data.append(instance)


        return data
    


if __name__ == "__main__":
    parser = InstanceSummaryParser("instances_data_sheet.xlsx")
    data = parser.parse()

    print("Dam data\n", data.get("dam")[:5])

    print("Irrigaiton\n", data.get("irrigation")[:5])
    print("Water Supply\n", data.get("water_supply")[:5])


