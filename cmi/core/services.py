from openpyxl import Workbook, load_workbook
from django.conf import settings
from django.core.files.uploadedfile import InMemoryUploadedFile
from django.contrib.auth.models import User

from io import BytesIO, StringIO
import os
import csv
from .models import Project, Engineer, Collector, DataInstance
from particular.models import Particular, ProjectType

from typing import Optional, List
from .utils.instance_summary_parser import InstanceSummaryParser

class ProjectDataTemplateGenerator:
    """
    Generates Excel and CSV templates for different data types.
    """
    def __init__(self, data_type=None):
        self.data_type = data_type
        self.headers = self._get_headers()
    
    def _get_headers(self):
        """
        Maps data types to their corresponding headers.
        """
        headers_map = {
            'project': [field.name for field in Project._meta.fields if field.name != 'id'],
            'engineer': [field.name for field in Engineer._meta.fields if field.name != 'id'],
            'collector': [field.name for field in Collector._meta.fields if field.name != 'id']
        }
        return headers_map.get(self.data_type, [])
    
    def create_template(self, file_format='xlsx'):
        """
        Creates an Excel or CSV template with predefined headers.
        Returns a file-like object (BytesIO).
        """
        if not self.headers:
            raise ValueError("Invalid data type provided.")

        if file_format == 'xlsx':
            return self._create_xlsx_template()
        elif file_format == 'csv':
            return self._create_csv_template()
        else:
            raise ValueError("Unsupported file format. Use 'xlsx' or 'csv'.")
    
    def _create_xlsx_template(self):
        """
        Internal method to create an XLSX template.
        """
        wb = Workbook()
        ws = wb.active
        ws.title = self.data_type.capitalize() + " Template"
        ws.append(self.headers)
        
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output
    
    def _create_csv_template(self):
        """
        Internal method to create a CSV template.
        """
        output = StringIO()
        writer = csv.writer(output)
        writer.writerow(self.headers)
        
        output.seek(0)
        return output
    

class ProjectDataParser:
    """
    Parses data from uploaded Excel or CSV files.
    """
    def __init__(self, data_file: InMemoryUploadedFile, data_type=None):
        self.data_file = data_file
        self.data_type = data_type

    def parse(self):
        """
        Determines the file type and calls the appropriate parser.
        """
        file_ext = self.data_file.name.split('.')[-1].lower()

        if file_ext == 'xlsx':
            return self._parse_xlsx()
        elif file_ext == 'csv':
            return self._parse_csv()
        else:
            raise ValueError("Unsupported file type. Please upload a .xlsx or .csv file.")

    def _parse_xlsx(self):
        """
        Internal method to parse an XLSX file.
        Returns a list of dictionaries, with each dictionary representing a row.
        """
        try:
            workbook = load_workbook(self.data_file)
            sheet = workbook.active
            rows = sheet.iter_rows(min_row=2, values_only=True)
            headers = [cell.value for cell in sheet[1]]
            
            data = []
            for row in rows:
                if any(row):  # Check if the row is not completely empty
                    row_data = dict(zip(headers, row))
                    data.append(row_data)
            return data
        except Exception as e:
            raise ValueError(f"Error parsing XLSX file: {e}")

    def _parse_csv(self):
        """
        Internal method to parse a CSV file.
        Returns a list of dictionaries.
        """
        try:
            # Decode the file in-memory
            decoded_file = self.data_file.read().decode('utf-8').splitlines()
            reader = csv.DictReader(decoded_file)
            
            data = [row for row in reader]
            return data
        except Exception as e:
            raise ValueError(f"Error parsing CSV file: {e}")
        


class InstanceDataMapper:
    def __init__(self, file: InMemoryUploadedFile, data_type=None, format="xlsx"):
        self.file = file
        self._data_type = data_type
        self.format = format


        if self.file is None:
            raise ValueError("Data is Empty")
    
       

    def save(self):
        try:
            parser = InstanceSummaryParser(self.file)
            parsed_data = parser.parse()

            self._import_instances(parsed_data)

        except Exception as e:
            raise e  

    def _import_instances(self, parsed_data):

        # dam_data = parsed_data.get("dam")
        # irrigation_data = parsed_data.get("irrigation")
        # ws_data = parsed_data.get("ws")
        
        
        for pr_type, instances in parsed_data.items():
            # foreign key lookups
            try:
                if pr_type == 'water_supply':
                    pr_type = pr_type.replace("_", " ")
                project_type = ProjectType.objects.get(name=pr_type.title())
            except ProjectType.DoesNotExist:
                raise ValueError(f"Project Type {pr_type} doesn't exist")
            
            print(f"Importing {pr_type} data instances....")

            for raw in instances:
                # check if the row is not void
                if raw.get('collector') is None  and  raw.get('PID') is None and raw.get("PE") is None :
                    continue

                # foreign key lookups
                try:
                    project = Project.objects.get(code=raw.get('project_code'))
                    # project_type = ProjectType.objects.get(name=pr_type.title())
                except Project.DoesNotExist:
                    raise ValueError(f"Project {raw.get('project_code')} not found")
                
                
                try:
                    engineer_name = raw.get("PE").split()
                    engineer_fname = engineer_name[0]
                    engineer=Engineer.objects.get(fname=engineer_fname)
                except Engineer.DoesNotExist:
                    raise ValueError(f"Engineer {raw.get('PE')} not found")
                    
                try:
                    cfname, cmname = raw['collector'].split()

                    # print(cfname, cmname)
                    collector = Collector.objects.get(fname=cfname)
                except Collector.DoesNotExist:
                    # raise ValueError("Data Collector not found")
                    # fname, mname = raw.get("collector").split(' ')
                    collector = Collector.objects.create(
                        engineer=engineer,
                        fname=cfname,
                        mname=cmname,
                        lname=cmname,
                    )
                    collector.projects.set([project])
                
                try:
                    pid = raw.get("PID")
                    if pid is not None:
                        particular = Particular.objects.get(pid=raw.get('PID'))
                    else:
                        continue
                except Particular.DoesNotExist:
                    raise ValueError("Particular doesnt exist")
                    

                try:
                    encoder = User.objects.get(username=raw.get("encoder"))
                except User.DoesNotExist:
                    # encoder = User.objects.create(
                    #     username=raw.get('encoder'),
                    #     email=None,
                    #     password=raw.get(encoder),
                    # )
                    encoder = None

                # boolean conversion
                encoded_flag = True if str(raw.get("encoded")).lower() in ["yes", "true", "1"] else False
                
                # check if instance already exists
                exists = DataInstance.objects.filter(
                    date=raw.get('date'),
                    particular=particular,
                    collector=collector,
                    project=project,
                )

               
                # if not exists create the instance
                if not exists:
                    instance = DataInstance.objects.create(
                        date=raw.get('date'),
                        project=project,
                        project_type=project_type,
                        collector=collector,
                        engineer=engineer,
                        particular=particular,
                        status=raw.get("status") or DataInstance.PENDING,
                        encoded=encoded_flag,
                        encoded_by=encoder,
                        review_comments=raw.get('problems'),
                    )
                else:
                    print("Instance already exists, skipping.")
                # create instance
                # instance = DataInstance.objects.create(
                #     date = raw.get('date'),
                #     project=project,
                #     project_type=project_type,
                #     collector=collector,
                #     particular=particular,
                #     status=raw.get("status") or DataInstance.PENDING,
                #     encoded=encoded_flag,
                #     encoded_by=encoder,
                #     review_comments = raw.get('problems'),
                # )
        print(f"Data parsed successfully {parsed_data.keys()}")
        
